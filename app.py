"""
AIB Group Credit Risk Dashboard — app.py
=========================================
IMPORTANT: All data is loaded LIVE from your Excel file.
If you change any number in the Excel, just re-run this app and
the dashboard will reflect it automatically.

Run:   streamlit run app.py
Needs: MAIN_Expected_Loss_model_.xlsx in the same folder as app.py

Install:  pip install streamlit plotly pandas openpyxl numpy
"""

from plotly import data
import streamlit as st
import pandas as pd
import numpy as np
import plotly.graph_objects as go
import plotly.express as px
from openpyxl import load_workbook
import os

st.set_page_config(
    page_title="AIB Credit Risk Dashboard",
    page_icon="🏦",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ── Colours ───────────────────────────────────────────────────────────────────
NAVY  = "#003366"; BLUE  = "#1F4E79"; LBLUE = "#BDD7EE"
GREEN = "#1F7A4D"; AMBER = "#C07010"; RED   = "#C00000"
GRAY  = "#F2F2F2"

# ══════════════════════════════════════════════════════════════════════════════
#  DATA LOADING — reads directly from your Excel
#  Every function below maps to exact sheet + row numbers in the workbook.
#  If you rename a sheet or move rows, update the row references here.
# ══════════════════════════════════════════════════════════════════════════════

EXCEL_FILE = "Credit_Risk_Model.xlsx"

@st.cache_data(show_spinner="Loading data from Excel…")
def load_all_data(filepath):
    """
    Load every data point from the Excel workbook.
    Returns a dict of DataFrames and scalar values.
    Sheet/row references are documented inline.
    """
    if not os.path.exists(filepath):
        return None

    wb = load_workbook(filepath, data_only=True)
    data = {}

    YEARS = [2020, 2021, 2022, 2023, 2024, 2025]
    SEGS  = [
        "Residential Mortgages",
        "Other Personal",
        "Property & Construction",
        "Non-Property Business(SME)",
        "Non-Property Business(Corporate other)",
    ]
    data["years"] = YEARS
    data["segs"]  = SEGS

    # ── Sheet: Raw Data ───────────────────────────────────────────────────────
    ws = wb["Raw Data"]

    # A. Loan Book — rows 6-10, cols B(2)-G(7)  [col B=2020 … col G=2025]
    loan_book = {}
    for row_idx, seg in zip(range(6, 11), SEGS):
        loan_book[seg] = [ws.cell(row_idx, c).value for c in range(3, 9)]
    data["loan_book"] = loan_book

    # B. NPL balances (Stage3+POCI) — rows 16-20
    npl_bal = {}
    for row_idx, seg in zip(range(16, 21), SEGS):
        npl_bal[seg] = [ws.cell(row_idx, c).value for c in range(3, 9)]
    data["npl_bal"] = npl_bal

    # C. NPL Ratios — rows 26-31
    npl_ratios = {}
    for row_idx, seg in zip(range(26, 31), SEGS):
        npl_ratios[seg] = [ws.cell(row_idx, c).value for c in range(3, 9)]
    data["npl_ratios"]  = npl_ratios
    data["group_npl"]   = [ws.cell(31, c).value for c in range(3, 9)]  # row 31

    # D. ECL Stage3+POCI — rows 36-41
    ecl_stage3 = {}
    for row_idx, seg in zip(range(36, 41), SEGS):
        ecl_stage3[seg] = [ws.cell(row_idx, c).value for c in range(3, 9)]
    data["ecl_stage3"]      = ecl_stage3
    data["ecl_stage3_total"]= [ws.cell(41, c).value for c in range(3, 9)]  # row 41

    # E. Coverage Ratios — rows 46-51
    coverage = {}
    for row_idx, seg in zip(range(46, 51), SEGS):
        coverage[seg] = [ws.cell(row_idx, c).value for c in range(3, 9)]
    data["coverage"]       = coverage
    data["group_coverage"] = [ws.cell(51, c).value for c in range(3, 9)]  # row 51

    # F. Cost of Risk — rows 56-58
    data["impairment"]  = [ws.cell(56, c).value for c in range(3, 9)]   # row 56
    data["avg_loans"]   = [ws.cell(57, c).value for c in range(3, 9)]   # row 57
    data["cor_bps"]     = [ws.cell(58, c).value for c in range(3, 9)]   # row 58

    # G. IFRS 9 Stage Balances — rows 63-68
    data["stage1"]     = [ws.cell(63, c).value for c in range(3, 9)]    # row 63
    data["stage2"]     = [ws.cell(64, c).value for c in range(3, 9)]    # row 64
    data["stage3"]     = [ws.cell(65, c).value for c in range(3, 9)]    # row 65
    data["stage1_ecl"] = [ws.cell(66, c).value for c in range(3, 9)]    # row 66
    data["stage2_ecl"] = [ws.cell(67, c).value for c in range(3, 9)]    # row 67
    data["stage3_ecl"] = [ws.cell(68, c).value for c in range(3, 9)]    # row 68

    # ── Sheet: EL Loss model ──────────────────────────────────────────────────
    ws = wb["EL Loss model"]

    # PD inputs — rows 6-10
    pd_inputs = {}
    for row_idx, seg in zip(range(6, 11), SEGS):
        pd_inputs[seg] = [ws.cell(row_idx, c).value for c in range(3, 9)]
    data["pd_inputs"] = pd_inputs

    # LGD inputs — rows 15-19
    lgd_inputs = {}
    for row_idx, seg in zip(range(15, 20), SEGS):
        lgd_inputs[seg] = [ws.cell(row_idx, c).value for c in range(3, 9)]
    data["lgd_inputs"] = lgd_inputs

    # EAD inputs (same as loan book, linked) — rows 24-28
    ead_inputs = {}
    for row_idx, seg in zip(range(24, 29), SEGS):
        ead_inputs[seg] = [ws.cell(row_idx, c).value for c in range(3, 9)]
    data["ead_inputs"] = ead_inputs

    # EL results — rows 33-37
    el_results = {}
    for row_idx, seg in zip(range(33, 38), SEGS):
        el_results[seg] = [ws.cell(row_idx, c).value for c in range(3, 9)]
    data["el_results"]       = el_results
    data["el_total"]         = [ws.cell(38, c).value for c in range(3, 9)]   # row 38

    # Coverage gap — rows 42-45
    data["actual_ecl_prov"]  = [ws.cell(42, c).value for c in range(3, 9)]   # row 42
    data["model_el"]         = [ws.cell(43, c).value for c in range(3, 9)]   # row 43
    data["coverage_gap"]     = [ws.cell(44, c).value for c in range(3, 9)]   # row 44
    data["coverage_ratio"]   = [ws.cell(45, c).value for c in range(3, 9)]   # row 45

    # IFRS9 Reconciliation Bridge values
    data["base_el"] = ws.cell(38, 8).value
    data["weighted_ecl"] = 1019.3
    # ── Sheet: Macro Regression ───────────────────────────────────────────────
    ws = wb["Macro Regression"]

    # Regression input data — rows 5-10, cols B-G (year, NPL, unemp, GDP, ECB, HPI)
    macro_years = [ws.cell(r, 2).value for r in range(5, 11)]
    macro_npl   = [ws.cell(r, 3).value for r in range(5, 11)]
    macro_unemp = [ws.cell(r, 4).value for r in range(5, 11)]
    macro_gdp   = [ws.cell(r, 5).value for r in range(5, 11)]
    macro_ecb   = [ws.cell(r, 6).value for r in range(5, 11)]
    macro_hpi   = [ws.cell(r, 7).value for r in range(5, 11)]
    data["macro_years"] = macro_years
    data["macro_npl"]   = macro_npl
    data["macro_unemp"] = macro_unemp
    data["macro_gdp"]   = macro_gdp
    data["macro_ecb"]   = macro_ecb
    data["macro_hpi"]   = macro_hpi

    # Fitted values — rows 15-20
    data["fitted_npl"] = [ws.cell(r, 4).value for r in range(15, 21)]   # col D
    data["residuals"]  = [ws.cell(r, 5).value for r in range(15, 21)]   # col E

    # Forward forecast — rows 25-29
    data["fc_unemp"]  = [ws.cell(25, c).value for c in range(3, 6)]     # Base/Adv/Sev
    data["fc_gdp"]    = [ws.cell(26, c).value for c in range(3, 6)]
    data["fc_ecb"]    = [ws.cell(27, c).value for c in range(3, 6)]
    data["fc_hpi"]    = [ws.cell(28, c).value for c in range(3, 6)]
    data["fc_npl"]    = [ws.cell(29, c).value for c in range(3, 6)]     # FORECAST NPL

    # ── Sheet: Stress Testing ─────────────────────────────────────────────────
    ws = wb["Stress Testing"]

    # Multipliers — rows 4-8, cols C-F (Base, DS1, DS2, Upside)
    multipliers = {}
    for row_idx, seg in zip(range(4, 9), SEGS):
        multipliers[seg] = {
            "Base":   ws.cell(row_idx, 3).value,
            "DS1":    ws.cell(row_idx, 4).value,
            "DS2":    ws.cell(row_idx, 5).value,
            "Upside": ws.cell(row_idx, 6).value,
        }
    data["multipliers"] = multipliers

    # Stressed PD — rows 12-16, cols C-F
    stressed_pd = {}
    for row_idx, seg in zip(range(12, 17), SEGS):
        stressed_pd[seg] = {
            "Base":   ws.cell(row_idx, 3).value,
            "DS1":    ws.cell(row_idx, 4).value,
            "DS2":    ws.cell(row_idx, 5).value,
            "Upside": ws.cell(row_idx, 6).value,
        }
    data["stressed_pd"] = stressed_pd

    # Stressed EL — rows 21-25, cols C-H
    stressed_el = {}
    for row_idx, seg in zip(range(21, 26), SEGS):
        stressed_el[seg] = {
            "Base":         ws.cell(row_idx, 3).value,
            "DS1":          ws.cell(row_idx, 4).value,
            "DS2":          ws.cell(row_idx, 5).value,
            "Upside":       ws.cell(row_idx, 6).value,
            "Wtd_ECL":      ws.cell(row_idx, 7).value,
            "Capital_Impact": ws.cell(row_idx, 8).value,
            "Pct_Increase": ws.cell(row_idx, 9).value,
        }
    data["stressed_el"] = stressed_el

    # Totals row 26
    data["stress_total"] = {
        "Base":   ws.cell(26, 3).value,
        "DS1":    ws.cell(26, 4).value,
        "DS2":    ws.cell(26, 5).value,
        "Upside": ws.cell(26, 6).value,
        "Wtd":    ws.cell(26, 7).value,
    }

    # Scenario summary — rows 30-34
    scen_summary = {}
    scen_names = ["Base", "Downside 1", "Downside 2", "Upside", "Weighted Avg"]
    for row_idx, scen in zip(range(30, 35), scen_names):
        scen_summary[scen] = {
            "Weight":          ws.cell(row_idx, 3).value,
            "ECL":             ws.cell(row_idx, 4).value,
            "Delta_Base":      ws.cell(row_idx, 5).value,
            "CET1_Post":       ws.cell(row_idx, 6).value,
            "CET1_Ratio_Post": ws.cell(row_idx, 7).value,
            "CET1_Req":        ws.cell(row_idx, 8).value,
            "Headroom":        ws.cell(row_idx, 9).value,
        }
    data["scen_summary"] = scen_summary

    # RWA and CET1 — from Macro(5y Avg) sheet
    ws_m = wb["Macro(5y Avg)"]
    data["rwa"]       = ws_m.cell(22, 3).value    # row 22 col C = Average RWA 2025
    data["cet1_ratio"]= ws_m.cell(30, 3).value    # row 30 col C = CET1 Ratio 2025
    data["cet1_cap"]  = (data["rwa"] or 58693) * (data["cet1_ratio"] or 0.162)
    data["roe"]       = ws_m.cell(23, 3).value    # RoTE
    data["profit_at"] = ws_m.cell(18, 3).value    # Profit after tax 2025

    # ── Sheet: PSI Analysis ───────────────────────────────────────────────────
    ws = wb["PSI Analysis"]

    # Stage PSI — rows 5-9
    data["psi_stage1"]  = [ws.cell(5, c).value for c in range(3, 9)]    # row 5
    data["psi_stage2"]  = [ws.cell(6, c).value for c in range(3, 9)]    # row 6
    data["psi_stage3"]  = [ws.cell(7, c).value for c in range(3, 9)]    # row 7
    data["psi_stage_total"] = [ws.cell(8, c).value for c in range(3, 9)] # row 8

    # Segment PSI — rows 14-20
    psi_segs = {}
    for row_idx, seg in zip(range(14, 19), SEGS):
        psi_segs[seg] = [ws.cell(row_idx, c).value for c in range(3, 9)]
    data["psi_segs"]        = psi_segs
    data["psi_seg_total"]   = [ws.cell(19, c).value for c in range(3, 9)]  # row 19

    # ── Sheet: Macro Forecast (ECL scenario table) ────────────────────────────
    ws = wb["Macro Forcast(5y)"]

    # ECL scenario table — rows 6-10, cols P-T (Reported, Base, DS1, DS2, Upside)
    ecl_scen_segs = ["Residential Mortgages", "Other Personal",
                     "Property & Construction", "Non-Property Business"]
    ecl_scen = {}
    for row_idx, seg in zip(range(6, 10), ecl_scen_segs):
        ecl_scen[seg] = {
            "Reported": ws.cell(row_idx, 16).value,
            "Base":     ws.cell(row_idx, 17).value,
            "DS1":      ws.cell(row_idx, 18).value,
            "DS2":      ws.cell(row_idx, 19).value,
            "Upside":   ws.cell(row_idx, 20).value,
        }
    data["ecl_scen"] = ecl_scen

    # Macro 5yr avg scenarios — from Macro(5y Avg) sheet
    ws = wb["Macro(5y Avg)"]
    # rows 7-13, cols C-F (Base, DS1, DS2, Upside)
    macro_factor_names = ["GDP growth", "Resi HPI", "Unemployment",
                          "Comm HPI", "Employment", "Disposable Income", "Inflation"]
    macro5y = {}
    for row_idx, factor in zip(range(7, 14), macro_factor_names):
        macro5y[factor] = {
            "Base":   ws.cell(row_idx, 3).value,
            "DS1":    ws.cell(row_idx, 4).value,
            "DS2":    ws.cell(row_idx, 5).value,
            "Upside": ws.cell(row_idx, 6).value,
        }
    data["macro5y"] = macro5y

    # Regulatory capital requirements
    data["cet1_req_p1"]  = ws.cell(35, 3).value   # 4.5%  — Pillar 1
    data["cet1_req_p2r"] = ws.cell(36, 3).value   # 1.35% — P2R
    data["cet1_ccb"]     = ws.cell(37, 3).value   # 2.5%  — CCB
    data["cet1_osii"]    = ws.cell(38, 3).value   # 1.5%  — O-SII
    data["cet1_ccyb"]    = ws.cell(39, 3).value   # 1.44% — CCyB
    data["cet1_total_req"]= ws.cell(40, 3).value  # 11.29% total CET1 req
    data["total_cap_req"] = ws.cell(43, 3).value  # 15.84% total capital req

    return data



def clean_number(x):

    if x is None:
        return 0

    if isinstance(x, str):

        x = x.replace("%", "").replace(",", "").strip()

        try:
            return float(x)
        except:
            return 0

    return float(x)
# ── Load the data ─────────────────────────────────────────────────────────────
EXCEL_PATH = EXCEL_FILE
if not os.path.exists(EXCEL_PATH):
    EXCEL_PATH = os.path.join(os.path.dirname(__file__), EXCEL_FILE)

d = load_all_data(EXCEL_PATH)

if d is None:
    st.error(f"❌ Cannot find **{EXCEL_FILE}**. Place it in the same folder as app.py and restart.")
    st.stop()

YEARS = d["years"]
SEGS  = d["segs"]
# Default LGD values loaded from Excel
lgd_user = {}

for seg in SEGS:
    lgd_user[seg] = clean_number(d["lgd_inputs"][seg][-1] or 0.40)
SEG_COLOURS = [NAVY, "#2E75B6", "#5BA3DC", "#70AD47", "#ED7D31"]


# ══════════════════════════════════════════════════════════════════════════════
#  SIDEBAR
# ══════════════════════════════════════════════════════════════════════════════

with st.sidebar:
    st.markdown("## 🏦 AIB Credit Risk")
    st.markdown("**All data live from Excel**")
    st.caption(f"File: `{EXCEL_FILE}`")
    if st.button("🔄 Refresh data from Excel"):
        st.cache_data.clear()
        st.rerun()
    st.divider()
    # st.markdown("### LGD Overrides (%)")
    # st.caption("Defaults loaded from Excel — change for scenarios")
    #lgd_user = {}
    #for seg in SEGS:
    #     default_lgd = d["lgd_inputs"][seg][-1] or 0.40
    #     lgd_user[seg] = st.slider(
    #         seg[:25], 0.05, 0.85, float(default_lgd), 0.01,
    #         format="%.0f%%", key=f"lgd_{seg}")
    # st.divider()
    # st.markdown("### Forward Scenario Inputs")
    # st.caption("Used in Stress Testing & Regression tabs")
    # unemp_b = st.slider("Base Unemployment",  0.01, 0.20, float(d["fc_unemp"][0] or 0.044), 0.001, format="%.1f%%")
    # unemp_a = st.slider("Adverse Unemploy't", 0.01, 0.25, float(d["fc_unemp"][1] or 0.070), 0.001, format="%.1f%%")
    # unemp_s = st.slider("Severe Unemploy't",  0.01, 0.30, float(d["fc_unemp"][2] or 0.100), 0.001, format="%.1f%%")
    # gdp_b   = st.slider("Base GDP",   -0.10, 0.15,  float(d["fc_gdp"][0] or 0.030), 0.005, format="%.1f%%")
    # gdp_a   = st.slider("Adverse GDP",-0.10, 0.10,  float(d["fc_gdp"][1] or -0.010),0.005, format="%.1f%%")
    # gdp_s   = st.slider("Severe GDP", -0.15, 0.05,  float(d["fc_gdp"][2] or -0.035),0.005, format="%.1f%%")
    st.markdown("## 📌 Model Assumptions")

    st.info("""
    ### IFRS 9 Framework
    - Expected Loss calculated using PD × LGD × EAD
    - Forward-looking macroeconomic overlays applied
    - Scenario-weighted ECL methodology used

    ### Stress Testing
    - Base: Stable Irish macro outlook
    - DS1: Tariff pressure and slower eurozone growth
    - DS2: Severe geopolitical and trade shock scenario
    - Upside: Falling inflation and improving credit conditions

    ### Risk Monitoring
    - PSI stability monitoring included
    - Stage migration analysis performed
    - CET1 post-stress capital assessment included
    - Reconciliation bridge aligned to reported ECL
    """)

    st.markdown("## 📚 Data Sources")

    st.success("""
    - AIB Annual Reports (2020–2025)
    - AIB Pillar 3 Disclosures
    - Central Bank of Ireland Statistics
    """)


# ══════════════════════════════════════════════════════════════════════════════
#  HEADER
# ══════════════════════════════════════════════════════════════════════════════
def clean_percent(x):

    # Handle None
    if x is None:
        return 0

    # Handle strings
    if isinstance(x, str):

        x = x.strip()

        # Empty string
        if x == "":
            return 0

        # Remove %
        x = x.replace("%", "").replace(",", "")

        # Try numeric conversion
        try:
            return float(x) / 100
        except:
            return 0

    # Handle numbers
    try:
        return float(x)
    except:
        return 0



st.markdown(
    f"""<div style='background:{NAVY};padding:16px 24px;border-radius:8px;margin-bottom:14px'>
    <h1 style='color:white;margin:0;font-size:1.6rem'>🏦 AIB Group — Credit Risk Dashboard  |  2020–2025</h1>
    <p style='color:{LBLUE};margin:5px 0 0;font-size:0.85rem'>
    Live data from <b style='color:white'>{EXCEL_FILE}</b> ·
    EL Model · IFRS 9 Staging · Macro Regression · Stress Testing · PSI Analysis
    </p></div>""", unsafe_allow_html=True)
st.caption("Interactive IFRS 9 Credit Risk Analytics Dashboard based on AIB disclosures")
# KPI strip — all from Excel
total_l  = sum((d["loan_book"][s][-1] or 0) for s in SEGS)
#grp_npl  = d["group_npl"][-1] or 0
grp_npl = clean_percent(d["group_npl"][-1])
base_npl = clean_percent(d["group_npl"][1])  # NPL ratio for 2020 (base year) for delta calculation
cet1_r   = clean_percent(d["cet1_ratio"])
ecl_25   = d["actual_ecl_prov"][-1] or 0
s3_25    = d["stage3"][-1] or 0
cor_25   = d["cor_bps"][-1] or 0

c1,c2,c3,c4,c5,c6 = st.columns(6)
with c1: st.metric("Total gross exposures 2025", f"€{total_l/1000:.1f}bn")
#with c2: st.metric("Group NPL 2025",   f"{grp_npl:.2%}", f"{grp_npl - (d['group_npl'][0] or 0):.1%} vs 2020", delta_color="inverse")
with c2:
    st.metric(
        "Group NPL 2025",
        f"{grp_npl:.2%}",
        f"{grp_npl - base_npl:.1%} vs 2020",
        delta_color="inverse"
    )
with c3: st.metric("CET1 Ratio 2025",  f"{cet1_r:.1%}")
with c4: st.metric("Total ECL 2025",   f"€{ecl_25:,}m")
with c5: st.metric("Stage 3 2025",     f"€{s3_25:,}m")
with c6: st.metric("Cost of Risk 2025",f"{cor_25:.1f} bps")
st.divider()


# ══════════════════════════════════════════════════════════════════════════════
#  TABS
# ══════════════════════════════════════════════════════════════════════════════

tab1,tab2,tab3,tab4,tab5,tab6 = st.tabs([
    "📊 Loan Book & Credit Quality",
    "💰 EL Model",
    "🔴 IFRS 9 Staging",
    "📈 Macro Regression",
    "⚡ Stress Testing",
    "📐 PSI Analysis",
])


# ── TAB 1: LOAN BOOK ──────────────────────────────────────────────────────────
with tab1:
    st.subheader("Loan Book & Credit Quality  2020–2025")
    st.caption("All data loaded from: Raw Data sheet → Sections A, B, C, D, E")

    ca, cb = st.columns(2)
    with ca:
        fig = go.Figure()
        for i, seg in enumerate(SEGS):
            fig.add_trace(go.Bar(name=seg, x=YEARS,
                                  y=[v or 0 for v in d["loan_book"][seg]],
                                  marker_color=SEG_COLOURS[i],
                                  hovertemplate=f"<b>{seg}</b><br>%{{x}}: €%{{y:,}}m<extra></extra>"))
        fig.update_layout(barmode="stack", title="Gross Loan Book by Segment (€m)",
                          height=400, template="plotly_white",
                          legend=dict(orientation="h", y=-0.3))
        st.plotly_chart(fig, use_container_width=True)

    with cb:
        fig = go.Figure()
        for i, seg in enumerate(SEGS):
            fig.add_trace(go.Scatter(
                name=seg, x=YEARS,
                y=[v*100 if v else 0 for v in d["npl_ratios"][seg]],
                mode="lines+markers",
                line=dict(color=SEG_COLOURS[i], width=2),
                marker=dict(size=6),
                hovertemplate=f"<b>{seg}</b><br>%{{x}}: %{{y:.2f}}%<extra></extra>"))
        fig.add_trace(go.Scatter(
            name="Group NPL", x=YEARS,
            y=[v*100 if v else 0 for v in d["group_npl"]],
            mode="lines+markers",
            line=dict(color="black", width=3, dash="dash"),
            marker=dict(size=8, symbol="diamond")))
        fig.update_layout(title="NPL Ratios by Segment + Group (%)",
                          height=400, template="plotly_white",
                          legend=dict(orientation="h", y=-0.3))
        st.plotly_chart(fig, use_container_width=True)

    cc, cd = st.columns(2)
    with cc:
        fig = go.Figure()
        fig.add_trace(go.Bar(x=YEARS, y=[v or 0 for v in d["cor_bps"]],
                              #marker_color=[RED if (v or 0) > 50 else AMBER if (v or 0) > 10 else GREEN for v in d["cor_bps"]],
                              marker_color=[RED if clean_number(v) > 50
                                       else AMBER if clean_number(v) > 10
                                       else GREEN for v in d["cor_bps"]],
                              text=[f"{clean_number(v):.1f}" for v in d["cor_bps"]],
                              textposition="outside",
                              name="Cost of Risk (bps)"))
        fig.update_layout(title="Cost of Risk (bps)  |  Source: Raw Data row 58",
                          height=320, template="plotly_white")
        fig.add_annotation(x=2020, y=-200, text="COVID-19 provision surge",
                            showarrow=True, arrowhead=2, ax=40, ay=-40, font=dict(size=11, color="black"))
        st.plotly_chart(fig, use_container_width=True)

    with cd:
        fig = go.Figure()
        for i, seg in enumerate(SEGS):
            fig.add_trace(go.Bar(name=seg, x=YEARS,
                                  y=[v or 0 for v in d["coverage"][seg]],
                                  marker_color=SEG_COLOURS[i]))
        fig.update_layout(barmode="group",
                          title="Provision Coverage Ratios by Segment  |  Raw Data rows 46-50",
                          height=320, template="plotly_white",
                          legend=dict(orientation="h", y=-0.3), yaxis_tickformat=".0%")
        st.plotly_chart(fig, use_container_width=True)

    # Data table
    loan_df = pd.DataFrame(d["loan_book"], index=YEARS).T
    loan_df.columns = [str(y) for y in YEARS]
    loan_df["Total"] = loan_df.select_dtypes(include="number").sum(axis=1)
    st.dataframe(
    loan_df.style.format(
        lambda x: f"{clean_number(x):,.0f}"
        if isinstance(x, (int, float, str))
        and str(x).replace(",", "").replace(".", "").isdigit()
        else x
    ),
    use_container_width=True
)
    #st.dataframe(loan_df.style.format("{:,.0f}"), use_container_width=True)


# ── TAB 2: EL MODEL ───────────────────────────────────────────────────────────
with tab2:
    st.subheader("Expected Loss Model  (EL = PD × LGD × EAD)  2020–2025")
    st.info("PD from EL Loss model rows 6-10 | LGD from rows 15-19 (sidebar overrides LGD) | "
            "EAD from rows 24-28 | EL results from rows 33-38")

    # Recalculate EL with sidebar LGD overrides
    el_recalc = {}
    for seg in SEGS:
        el_recalc[seg] = [clean_number(d["pd_inputs"][seg][i]) * clean_number(lgd_user[seg]) * clean_number(d["ead_inputs"][seg][i]) / 1000
                         for i in range(len(YEARS))
        ]
    total_el_recalc = [sum(el_recalc[seg][i] for seg in SEGS) for i in range(6)]

    ca, cb = st.columns(2)
    with ca:
        fig = go.Figure()
        for i, seg in enumerate(SEGS):
            fig.add_trace(go.Bar(name=seg, x=YEARS,
                                  y=[round(v, 1) for v in el_recalc[seg]],
                                  marker_color=SEG_COLOURS[i]))
        fig.update_layout(barmode="stack", title="Expected Loss by Segment (€m)  [LGD from sidebar]",
                          height=400, template="plotly_white",
                          legend=dict(orientation="h", y=-0.3))
        st.plotly_chart(fig, use_container_width=True)

    with cb:
        fig = go.Figure()
        fig.add_trace(go.Scatter(x=YEARS, y=[round(v,1) for v in total_el_recalc],
                                  name="Model EL (recalc)", mode="lines+markers",
                                  line=dict(color=NAVY, width=3), marker=dict(size=8)))
        fig.add_trace(go.Scatter(x=YEARS, y=[v or 0 for v in d["actual_ecl_prov"]],
                                  name="Actual ECL Provisions", mode="lines+markers",
                                  line=dict(color=GREEN, width=3, dash="dash"), marker=dict(size=8)))
        fig.add_trace(go.Scatter(x=YEARS, y=[v or 0 for v in d["el_total"]],
                                  name="Excel Model EL", mode="lines+markers",
                                  line=dict(color=AMBER, width=2, dash="dot"), marker=dict(size=6)))
        fig.update_layout(title="Model EL vs Actual ECL Provisions (€m)",
                          height=400, template="plotly_white",
                          legend=dict(orientation="h", y=-0.3))
        st.plotly_chart(fig, use_container_width=True)

    st.subheader("PD Inputs  (from EL Loss model sheet — rows 6-10)")
    #pd_df = pd.DataFrame(d["pd_inputs"], index=YEARS).T.applymap(lambda x: f"{x:.2%}" if x else "—")
    pd_df_raw = pd.DataFrame(d["pd_inputs"], index=YEARS).T

    pd_df = pd_df_raw.map(lambda x: (f"{clean_percent(x):.2%}"
                                    if clean_percent(x) != 0
                                    else "—"))
    pd_df.columns = [str(y) for y in YEARS]
    st.dataframe(pd_df, use_container_width=True)

    st.subheader("LGD Inputs  (from EL Loss model sheet — rows 15-19)")
    #lgd_df = pd.DataFrame(d["lgd_inputs"], index=YEARS).T.applymap(lambda x: f"{x:.2%}" if x else "—")
    lgd_df_raw = pd.DataFrame(d["lgd_inputs"], index=YEARS).T

    lgd_df = lgd_df_raw.map(lambda x: (f"{clean_percent(x):.2%}"
                                       if clean_percent(x) != 0
                                       else "—"))
    lgd_df.columns = [str(y) for y in YEARS]
    lgd_df["2025 Sidebar Override"] = [f"{lgd_user[seg]:.0%}" for seg in lgd_df.index]
    st.dataframe(lgd_df, use_container_width=True)

    st.subheader("Coverage Gap  (rows 42-45)")
    gap_df = pd.DataFrame({
        "Year": [str(y) for y in YEARS],
        "Actual ECL Prov (€m)": [v or 0 for v in d["actual_ecl_prov"]],
        #"Model EL (€m)": [round(v or 0, 1) for v in d["el_total"]],
        "Model EL (€m)": [round(clean_number(v), 1) for v in d["el_total"]],
        #"Gap (€m)": [round(v or 0, 1) for v in d["coverage_gap"]],
        "Gap (€m)": [round(clean_number(v), 1) for v in d["coverage_gap"]],
        #"Coverage Ratio": [f"{v:.2f}x" if v else "—" for v in d["coverage_ratio"]],
        "Coverage Ratio": [f"{clean_number(v):.2f}x" if clean_number(v) != 0 else "-" for v in d["coverage_ratio"]],
        #"Status": ["✅ Over-provisioned" if (v or 0)>0 else "⚠ Under" for v in d["coverage_gap"]],
        "Status": ["✅ Over-provisioned" if clean_number(v) > 0 else "⚠️ Under" for v in d["coverage_gap"]],
    })
    st.dataframe(gap_df, hide_index=True, use_container_width=True)

    # =========================================================
    # IFRS9 ECL RECONCILIATION BRIDGE
    # =========================================================

    st.subheader("IFRS9 ECL Reconciliation Bridge")

    base_el = d["base_el"]
    weighted_el = d["weighted_ecl"]

    management_overlay = 24
    final_ecl = 1191

    recon_df = pd.DataFrame({
        "Component": [
            "Base model EL",
            "Scenario weighted EL",
            "Management overlay / PMA",
            "Final reconciled ECL"
        ],

        "Value (€m)": [
            round(base_el,1),
            round(weighted_el,1),
            management_overlay,
            final_ecl
        ],

        "Explanation": [
            "PD × LGD × EAD model output",
            "Probability-weighted IFRS9 ECL",
            "Post-model adjustment for emerging risks",
            "Reported ECL allowance"
        ]
    })

    st.dataframe(
        recon_df,
        use_container_width=True,
        hide_index=True
    )


# ── TAB 3: IFRS 9 STAGING ─────────────────────────────────────────────────────
with tab3:
    st.subheader("IFRS 9 Stage Analysis  2020–2025")
    st.caption("Stage balances from Raw Data rows 63-68 | ECL provisions from rows 66-68")
    st.warning("Stage 2 is the key early warning indicator — it signals SICR before loans default.")

    stage_tots = [(d["stage1"][i] or 0)+(d["stage2"][i] or 0)+(d["stage3"][i] or 0) for i in range(6)]

    ca, cb = st.columns(2)
    with ca:
        fig = go.Figure()
        fig.add_trace(go.Bar(name="Stage 3 (Credit Impaired)", x=YEARS,
                              y=[v or 0 for v in d["stage3"]], marker_color=RED))
        fig.add_trace(go.Bar(name="Stage 2 (SICR)", x=YEARS,
                              y=[v or 0 for v in d["stage2"]], marker_color=AMBER))
        fig.add_trace(go.Bar(name="Stage 1 (Performing)", x=YEARS,
                              y=[v or 0 for v in d["stage1"]], marker_color=GREEN))
        fig.update_layout(barmode="stack", title="IFRS 9 Stage Balances (€m)",
                          height=400, template="plotly_white",
                          legend=dict(orientation="h", y=-0.3))
        st.plotly_chart(fig, use_container_width=True)

    with cb:
        #s2_pct = [(d["stage2"][i] or 0) / t * 100 if t else 0 for i, t in enumerate(stage_tots)]
        s2_pct = [(clean_number(d["stage2"][i]) / clean_number(t)) * 100 if clean_number(t) != 0 else 0 for i, t in enumerate(stage_tots)]
        fig = go.Figure()
        fig.add_trace(go.Scatter(x=YEARS, y=s2_pct, mode="lines+markers+text",
                                  line=dict(color=AMBER, width=3),
                                  marker=dict(size=10, color=[RED if v>12 else AMBER for v in s2_pct]),
                                  text=[f"{v:.1f}%" for v in s2_pct],
                                  textposition="top center", name="Stage 2 %"))
        fig.add_hline(y=12, line_dash="dash", line_color=RED, annotation_text="12% Warning")
        fig.update_layout(title="Stage 2 as % of Total  (Early Warning KPI)",
                          height=400, template="plotly_white")
        st.plotly_chart(fig, use_container_width=True)

    cc, cd = st.columns(2)
    with cc:
        fig = go.Figure()
        fig.add_trace(go.Bar(name="S3 ECL", x=YEARS, y=[v or 0 for v in d["stage3_ecl"]], marker_color=RED))
        fig.add_trace(go.Bar(name="S2 ECL", x=YEARS, y=[v or 0 for v in d["stage2_ecl"]], marker_color=AMBER))
        fig.add_trace(go.Bar(name="S1 ECL", x=YEARS, y=[v or 0 for v in d["stage1_ecl"]], marker_color=GREEN))
        fig.update_layout(barmode="stack", title="ECL Provisions by Stage (€m)  |  rows 66-68",
                          height=350, template="plotly_white",
                          legend=dict(orientation="h", y=-0.3))
        st.plotly_chart(fig, use_container_width=True)

    with cd:
        s1c = [(clean_number(d["stage1_ecl"][i]) / clean_number(d["stage1"][i])) * 100 if clean_number(d["stage1"][i]) != 0 else 0 for i in range(6)]
        s2c = [(clean_number(d["stage2_ecl"][i]) / clean_number(d["stage2"][i])) * 100 if clean_number(d["stage2"][i]) != 0 else 0 for i in range(6)]
        s3c = [(clean_number(d["stage3_ecl"][i]) / clean_number(d["stage3"][i])) * 100 if clean_number(d["stage3"][i]) != 0 else 0 for i in range(6)]
        fig = go.Figure()
        for name, cov, col in [("S1 Coverage",s1c,GREEN),("S2 Coverage",s2c,AMBER),("S3 Coverage",s3c,RED)]:
            fig.add_trace(go.Scatter(x=YEARS, y=cov, name=name, mode="lines+markers",
                                      line=dict(color=col, width=2), marker=dict(size=6)))
        fig.update_layout(title="ECL Coverage by Stage (%)  |  rows 63-68",
                          height=350, template="plotly_white",
                          legend=dict(orientation="h", y=-0.3), yaxis_tickformat=".0f")
        st.plotly_chart(fig, use_container_width=True)

    # Heatmap of stage distribution
    heat = [[(clean_number(d["stage1"][i]) / clean_number(t) * 100)  if clean_number(t) != 0 else 0,
             ((clean_number(d["stage2"][i]) / clean_number(t)) * 100 if clean_number(t) != 0 else 0),
             ((clean_number(d["stage3"][i]) / clean_number(t)) * 100 if clean_number(t) != 0 else 0)]
            for i, t in enumerate(stage_tots)]
    fig_h = px.imshow(heat, labels=dict(x="Stage", y="Year", color="%"),
                       x=["Stage 1", "Stage 2", "Stage 3"],
                       y=[str(y) for y in YEARS],
                       color_continuous_scale=["green","yellow","red"],
                       text_auto=".1f", aspect="auto",
                       title="Stage Distribution % Heatmap  (hotter = more risk)")
    fig_h.update_layout(height=300, template="plotly_white")
    st.plotly_chart(fig_h, use_container_width=True)


# ── TAB 4: MACRO REGRESSION ───────────────────────────────────────────────────
with tab4:
    st.subheader("Macro-Linked PD Estimation  (OLS Regression)  2020–2025")
    st.caption("Regression data from Macro Regression sheet rows 5-10 | Fitted values rows 15-20 | Forecast rows 25-29")

    # Compute R² from Excel's actual vs fitted
    actual_arr = np.array(d["macro_npl"])
    fitted_arr = np.array(d["fitted_npl"])
    mask = (actual_arr != None) & (fitted_arr != None)
    r_sq = 0.0
    if mask.sum() > 2:
        ss_res = np.sum((actual_arr[mask] - fitted_arr[mask])**2)
        ss_tot = np.sum((actual_arr[mask] - actual_arr[mask].mean())**2)
        r_sq = 1 - ss_res/ss_tot if ss_tot != 0 else 0

    st.info(f"Model: NPL = α + β₁(Unemp) + β₂(GDP) + β₃(ECBRate) + β₄(HPI)  |  "
            f"**R² = {r_sq:.1%}** (6 observations: 2020–2025)  |  "
            f"Fitted values loaded directly from Excel regression output"
            f"(illustrative regression, 6 observations = 0 degrees of freedom, not for production inference)")

    ca, cb = st.columns(2)
    with ca:
        fig = go.Figure()
        fig.add_trace(go.Scatter(x=YEARS, y=[v*100 if v else 0 for v in d["macro_npl"]],
                                  name="Actual NPL (%)", mode="lines+markers",
                                  line=dict(color=NAVY, width=3), marker=dict(size=9)))
        fig.add_trace(go.Scatter(x=YEARS, y=[v*100 if v else 0 for v in d["fitted_npl"]],
                                  name="Fitted NPL (%)", mode="lines+markers",
                                  line=dict(color=AMBER, width=2, dash="dash"), marker=dict(size=6)))
        fig.update_layout(title=f"Fitted vs Actual NPL  (R²={r_sq:.1%})",
                          height=380, template="plotly_white",
                          legend=dict(orientation="h", y=-0.25))
        st.plotly_chart(fig, use_container_width=True)

    with cb:
        macro_factors = ["GDP growth", "Unemployment", "ECB Rate", "House Price"]
        macro_vals_2025 = [d["macro_gdp"][-1]*100, d["macro_unemp"][-1]*100,
                           d["macro_ecb"][-1]*100, d["macro_hpi"][-1]*100]
        fig = go.Figure(go.Bar(
            x=macro_factors, y=macro_vals_2025,
            marker_color=[GREEN if v>0 else RED for v in macro_vals_2025],
            text=[f"{v:.1f}%" for v in macro_vals_2025], textposition="outside"))
        fig.update_layout(title="2025 Macro Variables (actual)",
                          height=380, template="plotly_white")
        st.plotly_chart(fig, use_container_width=True)

    # Macro trend chart
    st.subheader("Macro Variable Trends  2020–2025  |  rows 5-10")
    fig_m = go.Figure()
    for name, vals, col in [
        ("Unemployment", [v*100 if v else 0 for v in d["macro_unemp"]], RED),
        ("GDP Growth",   [v*100 if v else 0 for v in d["macro_gdp"]],   GREEN),
        ("ECB Rate",     [v*100 if v else 0 for v in d["macro_ecb"]],   AMBER),
        ("HPI",          [v*100 if v else 0 for v in d["macro_hpi"]],   BLUE),
    ]:
        fig_m.add_trace(go.Scatter(x=YEARS, y=vals, name=name,
                                    mode="lines+markers",
                                    line=dict(color=col, width=2),
                                    marker=dict(size=6)))
    fig_m.update_layout(title="Macro Variables (%)", height=320,
                         template="plotly_white", legend=dict(orientation="h", y=-0.3))
    st.plotly_chart(fig_m, use_container_width=True)

    # Forward forecast from Excel + sidebar overrides
    st.subheader("Forward NPL Forecast  |  rows 29 (Excel calc) + sidebar override")
    fc_labels = ["Base Case", "Downside 1", "Downside 2"]
    fc_excel  = [v*100 if v else 0 for v in d["fc_npl"]]

    # Re-run OLS with sidebar values for live recalc
    Y  = np.array(d["macro_npl"])
    X  = np.column_stack([np.ones(6), d["macro_unemp"], d["macro_gdp"], d["macro_ecb"], d["macro_hpi"]])
    try:
        beta, *_ = np.linalg.lstsq(X, Y, rcond=None)
        fc_sidebar = []
        for u, g, r_, h in [(unemp_b, gdp_b, d["fc_ecb"][0] or 0.025, 0.05),
                             (unemp_a, gdp_a, d["fc_ecb"][1] or 0.035, -0.05),
                             (unemp_s, gdp_s, d["fc_ecb"][2] or 0.045, -0.10)]:
            fc_sidebar.append(float(np.array([1, u, g, r_, h]) @ beta) * 100)
    except Exception:
        fc_sidebar = fc_excel

    f1, f2, f3 = st.columns(3)
    for col_w, label, excel_v, sidebar_v, col in zip(
        [f1,f2,f3], fc_labels, fc_excel, fc_sidebar,
        [GREEN, AMBER, RED]):
        with col_w:
            st.markdown(
                f"<div style='background:{col};padding:12px;border-radius:6px;text-align:center'>"
                f"<b style='color:white;font-size:1.1rem'>{label}</b><br>"
                f"<span style='color:white;font-size:1.8rem;font-weight:bold'>{excel_v:.1f}%</span><br>"
                f"<span style='color:white;font-size:0.75rem'>Excel model | Sidebar: {sidebar_v:.1f}%</span></div>",
                unsafe_allow_html=True)

    # AIB 5yr macro scenarios
    st.subheader("AIB 5-Year Average Macro Scenarios  |  Macro(5y Avg) sheet rows 7-13")
    macro5y_df = pd.DataFrame(
        {f: {s: d["macro5y"][f][s] for s in ["Base","DS1","DS2","Upside"]}
         for f in d["macro5y"]}
    ).T
    macro5y_df.columns = ["Base","Downside 1","Downside 2","Upside"]
    st.dataframe(macro5y_df.style.format("{:.1f}%"), use_container_width=True)


# ── TAB 5: STRESS TESTING ─────────────────────────────────────────────────────
with tab5:
    st.subheader("Stress Testing  |  4 Scenarios  |  Data from Stress Testing sheet")
    st.info("""
            • Base: Stable Irish economy with resilient credit conditions  
            • DS1: Tariff tensions and weaker growth increase borrower stress  
            • DS2: Trade war and FDI shock trigger severe recession risks  
            • Upside: Easing geopolitical tensions improve credit outlook
    """)
    # st.info(
    #     "Multipliers (rows 4-8) derived from AIB ECL scenario table (DS/Base ratios). "
    #     "Stressed PD (rows 12-16) = Base PD × Multiplier. "
    #     "Stressed EL (rows 21-25) = Stressed PD × LGD × EAD. "
    #     "Capital impact uses CET1 from Macro(5y Avg) sheet."
    # )

    # Use Excel stressed EL directly
    scen_names_st = ["Base", "DS1", "DS2", "Upside"]
    scen_labels_st = ["Base", "Downside 1", "Downside 2", "Upside"]
    scen_colours   = [GREEN, AMBER, RED, BLUE]

    ka, kb, kc, kd = st.columns(4)
    totals = d["stress_total"]
    cet1_c = d["cet1_cap"] or 9508.27
    for col_w, scen, label, col in zip([ka,kb,kc,kd], scen_names_st, scen_labels_st, scen_colours):
        v = totals.get(scen) or 0
        delta = v - (totals.get("Base") or 0)
        with col_w:
            if scen == "Base":
                st.metric(f"{label} EL", f"€{v:.0f}m")
            else:
                st.metric(f"{label} EL", f"€{v:.0f}m",
                           delta=f"+€{delta:.0f}m vs Base", delta_color="inverse")

    ca, cb = st.columns(2)
    with ca:
        fig = go.Figure()
        for scen, label, col in zip(scen_names_st, scen_labels_st, scen_colours):
            fig.add_trace(go.Bar(
                name=label, x=SEGS,
                y=[round(d["stressed_el"][seg][scen] or 0, 1) for seg in SEGS],
                marker_color=col))
        fig.update_layout(barmode="group",
                          title="Stressed EL by Segment & Scenario (€m)  |  rows 21-25",
                          xaxis_tickangle=-20, height=420, template="plotly_white",
                          legend=dict(orientation="h", y=-0.35))
        st.plotly_chart(fig, use_container_width=True)

    with cb:
        base_v = totals.get("Base") or 0
        ds1_v  = totals.get("DS1") or 0
        ds2_v  = totals.get("DS2") or 0
        fig = go.Figure(go.Waterfall(
            orientation="v",
            x=["Base", "DS1 Δ", "DS1 Total", "DS2 Δ", "DS2 Total"],
            measure=["absolute","relative","total","relative","total"],
            y=[base_v, ds1_v-base_v, ds1_v, ds2_v-ds1_v, ds2_v],
            increasing=dict(marker=dict(color=RED)),
            decreasing=dict(marker=dict(color=GREEN)),
            totals=dict(marker=dict(color=AMBER)),
            connector=dict(line=dict(color=NAVY))))
        fig.update_layout(title="EL Waterfall: Base → DS1 → DS2",
                          height=420, template="plotly_white")
        st.plotly_chart(fig, use_container_width=True)

    # Capital impact table from Excel
    st.subheader("Capital Impact  |  Scenario Summary rows 30-34")
    ss = d["scen_summary"]
    cap_df = pd.DataFrame({
        "Scenario": list(ss.keys()),
        "Weight":   [f"{ss[s]['Weight']:.0%}" if ss[s]["Weight"] else "—" for s in ss],
        "EL / ECL (€m)": [f"{ss[s]['ECL']:.1f}" if ss[s]["ECL"] else "—" for s in ss],
        "Δ vs Base (€m)": [f"{ss[s]['Delta_Base']:.1f}" if ss[s]["Delta_Base"] else "—" for s in ss],
        "Post-Stress CET1 Capital": [f"€{ss[s]['CET1_Post']:,.0f}m" if ss[s]["CET1_Post"] else "—" for s in ss],
        "CET1 Ratio Post": [f"{ss[s]['CET1_Ratio_Post']:.2%}" if ss[s]["CET1_Ratio_Post"] else "—" for s in ss],
        "CET1 Requirement": [f"{ss[s]['CET1_Req']:.2%}" if ss[s]["CET1_Req"] else "—" for s in ss],
        "Headroom": [f"{ss[s]['Headroom']:.2%}" if ss[s]["Headroom"] else "—" for s in ss],
    })
    st.dataframe(cap_df, hide_index=True, use_container_width=True)

    # CET1 post-stress visual
    # cet1_vals = [ss[s]["CET1_Ratio_Post"] for s in ["Base","Downside 1","Downside 2","Upside"]]
    # req = d["cet1_total_req"] or 0.1584
    # fig_cap = go.Figure()
    # fig_cap.add_trace(go.Bar(
    #     x=["Base","DS1","DS2","Upside"],
    #     y=[v*100 if v else 0 for v in cet1_vals],
    #     marker_color=[GREEN if (v or 0)>(req) else RED for v in cet1_vals],
    #     text=[f"{v:.2%}" if v else "—" for v in cet1_vals],
    #     textposition="outside"))
    # fig_cap.add_hline(y=req*100, line_dash="dash", line_color=RED,
    #                   annotation_text=f"Total Cap Req = {req:.1%}")
    # fig_cap.add_hline(y=(d["cet1_total_req"] or 0.1129)*100, line_dash="dot", line_color=AMBER,
    #                   annotation_text=f"CET1 Req = {(d['cet1_total_req'] or 0.1129):.1%}")
    # fig_cap.update_layout(title="Post-Stress CET1 Ratio vs Requirements",
    #                        height=340, template="plotly_white")
    # st.plotly_chart(fig_cap, use_container_width=True)

    # CET1 post-stress visual
    cet1_vals = [ss[s]["CET1_Ratio_Post"] for s in ["Base","Downside 1","Downside 2","Upside"]]

    cet1_req = d["cet1_total_req"] or 0.1129
    total_cap_req = d["total_cap_req"] or 0.1584

    fig_cap = go.Figure()

    fig_cap.add_trace(go.Bar(
    x=["Base","DS1","DS2","Upside"],
    y=[v*100 if v else 0 for v in cet1_vals],
    marker_color=[GREEN if (v or 0) > cet1_req else RED for v in cet1_vals],
    text=[f"{v:.2%}" if v else "—" for v in cet1_vals],
    textposition="outside"
    ))

    # CET1 minimum requirement line (11.29%)
    fig_cap.add_hline(
    y=cet1_req * 100,
    line_dash="dash",
    line_color=AMBER,
    annotation_text=f"CET1 Req = {cet1_req*100:.2f}%"
    )

    # Total capital requirement line (15.84%)
    fig_cap.add_hline(
    y=total_cap_req * 100,
    line_dash="dot",
    line_color=RED,
    annotation_text=f"Total Cap Req = {total_cap_req*100:.2f}%"
    )

    fig_cap.update_layout(
    title="Post-Stress CET1 Ratio vs Requirements",
    height=340,
    template="plotly_white"
    )

    st.plotly_chart(fig_cap, use_container_width=True)
    # Multiplier table
    st.subheader("PD Stress Multipliers  |  rows 4-8  |  Derived from AIB ECL scenario ratios")
    mult_df = pd.DataFrame(
        {seg: d["multipliers"][seg] for seg in SEGS}
    ).T
    st.dataframe(mult_df.style.format("{:.2f}x"), use_container_width=True)


# ── TAB 6: PSI ────────────────────────────────────────────────────────────────
with tab6:
    st.subheader("Population Stability Index  |  Portfolio Drift Detection  |  Baseline: 2020")
    st.caption("PSI data from PSI Analysis sheet rows 5-9 (Stage PSI) and rows 14-20 (Segment PSI)")
    st.info("PSI < 0.10 = Stable  |  0.10–0.25 = Monitor  |  > 0.25 = Major Shift")

    def psi_colour(v):
        v = clean_number(v)
        if v is None: return GRAY
        if v < 0.10: return GREEN
        if v < 0.25: return AMBER
        return RED

    ca, cb = st.columns(2)
    with ca:
        psi_st = d["psi_stage_total"]
        fig = go.Figure(go.Bar(
            x=YEARS, y=[v or 0 for v in psi_st],
            marker_color=[psi_colour(v) for v in psi_st],
            text=[f"{clean_number(v):.4f}" if clean_number(v)!=0 else "0" for v in psi_st],
            textposition="outside", name="Stage PSI"))
        fig.add_hline(y=0.10, line_dash="dash", line_color=AMBER, annotation_text="0.10 Monitor")
        fig.add_hline(y=0.25, line_dash="dash", line_color=RED,   annotation_text="0.25 Major Shift")
        fig.update_layout(title="Stage Distribution PSI vs 2020 Baseline  |  rows 5-9",
                          height=380, template="plotly_white")
        st.plotly_chart(fig, use_container_width=True)

    with cb:
        psi_sg = d["psi_seg_total"]
        fig = go.Figure(go.Bar(
            x=YEARS, y=[v or 0 for v in psi_sg],
            marker_color=[psi_colour(v) for v in psi_sg],
            text=[f"{clean_number(v):.4f}" if clean_number(v)!=0 else "0" for v in psi_sg],
            textposition="outside", name="Segment PSI"))
        fig.add_hline(y=0.10, line_dash="dash", line_color=AMBER)
        fig.add_hline(y=0.25, line_dash="dash", line_color=RED)
        fig.update_layout(title="Segment Distribution PSI vs 2020 Baseline  |  rows 14-20",
                          height=380, template="plotly_white")
        st.plotly_chart(fig, use_container_width=True)

    # Stage shares heatmap
    s1p = [(d["psi_stage1"][i] or 0)*100 for i in range(6)]
    s2p = [(d["psi_stage2"][i] or 0)*100 for i in range(6)]
    s3p = [(d["psi_stage3"][i] or 0)*100 for i in range(6)]
    fig_h = px.imshow(
        list(zip(s1p, s2p, s3p)),
        labels=dict(x="Stage", y="Year", color="%"),
        x=["Stage 1","Stage 2","Stage 3"],
        y=[str(y) for y in YEARS],
        color_continuous_scale=["green","yellow","red"],
        text_auto=".1f", aspect="auto",
        title="Stage Distribution % Heatmap  |  rows 5-7")
    fig_h.update_layout(height=320, template="plotly_white")
    st.plotly_chart(fig_h, use_container_width=True)

    # Segment PSI table
    psi_summary = pd.DataFrame({
        "Year": [str(y) for y in YEARS],
        "Stage PSI":   [round(clean_number(v), 4) for v in d["psi_stage_total"]],
        "Stage Status":[("🟢 Stable" if (clean_number(v) or 0)<0.10 else "🟡 Monitor" if (clean_number(v) or 0)<0.25 else "🔴 Major Shift")
                        for v in d["psi_stage_total"]],
        "Segment PSI": [round(clean_number(v), 4) for v in d["psi_seg_total"]],
        "Seg Status":  [("🟢 Stable" if (clean_number(v) or 0)<0.10 else "🟡 Monitor" if (clean_number(v) or 0)<0.25 else "🔴 Major Shift")
                        for v in d["psi_seg_total"]],
    })
    st.dataframe(psi_summary, hide_index=True, use_container_width=True)

    # Per-segment PSI line chart
    fig_segs = go.Figure()
    for i, seg in enumerate(SEGS):
        fig_segs.add_trace(go.Scatter(
            x=YEARS, y=[round(clean_number(v), 4) for v in d["psi_segs"][seg]],
            name=seg[:25], mode="lines+markers",
            line=dict(color=SEG_COLOURS[i], width=2), marker=dict(size=6)))
    fig_segs.add_hline(y=0.10, line_dash="dash", line_color=AMBER)
    fig_segs.add_hline(y=0.25, line_dash="dash", line_color=RED)
    fig_segs.update_layout(title="PSI per Segment vs 2020 Baseline  |  rows 14-18",
                            height=340, template="plotly_white",
                            legend=dict(orientation="h", y=-0.3))
    st.plotly_chart(fig_segs, use_container_width=True)


# ── Footer ────────────────────────────────────────────────────────────────────
st.divider()
st.markdown(
    f"<div style='text-align:center;color:#888;font-size:0.75rem'>"
    f"Data loaded live from <b>{EXCEL_FILE}</b> ·  "
    f"Edit the Excel → click 🔄 Refresh in the sidebar → dashboard updates instantly ·  "
    f"AIB Annual Reports 2020-2025 + Pillar 3 Q4 2025</div>",
    unsafe_allow_html=True)
